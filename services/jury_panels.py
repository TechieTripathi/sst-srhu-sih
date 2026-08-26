"""
Jury Panel Service — panel rosters, team assignment and scoring progress.

Panel membership is a scalar `panel_no` on the judges document and on the teams
document. There is deliberately no panel collection holding an array of team ids:
an array cannot structurally stop two panels claiming the same team, this
codebase has no transactions, and moving a team would be a non-atomic $pull plus
$push with a window where the team belongs to zero or two panels. A scalar is
single-valued by construction, which is exactly the invariant the event needs.
"""

import math
from datetime import datetime

from bson.objectid import ObjectId

from models.database import (
    get_evaluations_collection,
    get_judges_collection,
    get_teams_collection,
)
from services.audit import log_audit
from services.jury_scope import (
    FILTER_EXCEPTION,
    FILTER_GROUP,
    PANEL_COUNT,
    PANEL_NUMBERS,
    SCOPE_ALL,
    SCOPE_ASSIGNED,
    is_exception_jury,
    judge_panel_no,
    panel_judge_counts,
    judge_user_ids_by_panel,
)

DEFAULT_STAGE = 'final_presentation'


# --------------------------------------------------------------------------- #
# Reads
# --------------------------------------------------------------------------- #

def panel_teams(panel_no):
    return list(get_teams_collection().find({'panel_no': int(panel_no)}).sort('team_name', 1))


def panel_judges(panel_no):
    return list(
        get_judges_collection().find({'panel_no': int(panel_no), **FILTER_GROUP}).sort('name', 1)
    )


def exception_judges():
    return list(get_judges_collection().find(FILTER_EXCEPTION).sort('name', 1))


def unassigned_teams():
    """Teams with no panel. Only exception jury can score these.

    $nin also matches documents where the field is absent or null, so this one
    query covers never-assigned, cleared and malformed values alike.
    """
    return list(
        get_teams_collection()
        .find({'panel_no': {'$nin': list(PANEL_NUMBERS)}})
        .sort([('created_at', 1), ('_id', 1)])
    )


def list_panels(stage_id=DEFAULT_STAGE):
    """All five panels with their rosters, teams and scoring progress."""
    counts = panel_judge_counts()
    by_panel_users = judge_user_ids_by_panel()
    evaluations_col = get_evaluations_collection()

    panels = []
    for n in PANEL_NUMBERS:
        teams = panel_teams(n)
        team_ids = [str(t['_id']) for t in teams]
        user_ids = by_panel_users.get(n, [])

        expected = len(user_ids) * len(teams)
        submitted = 0
        if user_ids and team_ids:
            submitted = evaluations_col.count_documents({
                'judge_id': {'$in': user_ids},
                'team_id': {'$in': team_ids},
                'stage_id': stage_id,
                'status': 'submitted',
            })

        panels.append({
            'panel_no': n,
            'judge_count': counts.get(n, 0),
            'team_count': len(teams),
            'judges': panel_judges(n),
            'teams': teams,
            'submitted': submitted,
            'expected': expected,
            'pct': round(submitted / expected * 100) if expected else 0,
        })
    return panels


def panel_detail(panel_no, stage_id=DEFAULT_STAGE):
    for panel in list_panels(stage_id=stage_id):
        if panel['panel_no'] == int(panel_no):
            return panel
    return None


def roster_overview(stage_id=DEFAULT_STAGE):
    """Counts for the panels landing page and the admin dashboard tile."""
    teams_col = get_teams_collection()
    judges_col = get_judges_collection()

    total_teams = teams_col.count_documents({})
    assigned = teams_col.count_documents({'panel_no': {'$in': list(PANEL_NUMBERS)}})

    return {
        'judges': judges_col.count_documents({}),
        'exception': judges_col.count_documents(FILTER_EXCEPTION),
        'group': judges_col.count_documents(FILTER_GROUP),
        'total_teams': total_teams,
        'assigned': assigned,
        'unassigned': total_teams - assigned,
    }


def judge_submitted_count(judge, stage_id=DEFAULT_STAGE):
    """How many evaluations this judge has already submitted.

    Used to warn an admin before moving someone between panels.
    """
    user_id = (judge or {}).get('user_id')
    if not user_id:
        return 0
    return get_evaluations_collection().count_documents({
        'judge_id': user_id,
        'stage_id': stage_id,
        'status': 'submitted',
    })


def build_panels_workbook(stage_id=DEFAULT_STAGE):
    """Build the jury-panels Excel workbook and return it as a BytesIO.

    Seven tabs: Overview, Panel 1..5 (judge roster + assigned teams, printable
    per coordinator), and a flat All teams sheet for sorting and pivoting.
    Includes live scoring progress so the same file doubles as the mid-event
    tracking sheet.

    openpyxl is imported here rather than at module level so the app's cold
    start never pays for it - this module is on the judge-dashboard hot path.
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    from services.results_calculator import calculate_all_teams_scores

    panels = list_panels(stage_id=stage_id)
    overview = roster_overview()
    exceptions = exception_judges()
    unassigned = unassigned_teams()
    # One pass over all teams for the whole workbook, indexed by team_id -
    # never recomputed per row.
    scores_by_team = {r['team_id']: r for r in calculate_all_teams_scores(stage_id)}

    TITLE = Font(bold=True, size=14)
    H2 = Font(bold=True, size=11)
    HEADER = Font(bold=True)
    NOTE = Font(italic=True, color='666666')

    def write_header(ws, row, headers, widths=None):
        for col, text in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=text)
            cell.font = HEADER
        if widths:
            for col, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(col)].width = width
        return row + 1

    def team_progress(team):
        """(panel evals, expected, exception evals, status) for one team doc."""
        s = scores_by_team.get(str(team['_id']))
        if not s:
            return 0, 0, 0, 'NOT_SCORED'
        return s['group_count'], s['group_expected'], s['exception_count'], s['status']

    wb = Workbook()

    # ------------------------------------------------------------------ #
    # Overview
    # ------------------------------------------------------------------ #
    ws = wb.active
    ws.title = 'Overview'
    ws.cell(row=1, column=1, value='TechForge 3.0 — Jury Panels').font = TITLE
    ws.cell(row=2, column=1,
            value=f"Exported {datetime.utcnow().strftime('%d %b %Y %H:%M')} UTC · "
                  f"Stage: {stage_id}").font = NOTE

    row = 4
    ws.cell(row=row, column=1, value='PANEL SUMMARY').font = H2
    row = write_header(ws, row + 1,
                       ['Panel', 'Judges', 'Teams', 'Evaluations submitted',
                        'Expected', 'Progress %'],
                       widths=[10, 10, 10, 22, 12, 12])
    for p in panels:
        ws.cell(row=row, column=1, value=f"Panel {p['panel_no']}")
        ws.cell(row=row, column=2, value=p['judge_count'])
        ws.cell(row=row, column=3, value=p['team_count'])
        ws.cell(row=row, column=4, value=p['submitted'])
        ws.cell(row=row, column=5, value=p['expected'])
        ws.cell(row=row, column=6, value=p['pct'])
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='EXCEPTION JURY').font = H2
    ws.cell(row=row + 1, column=1,
            value='Score every team, carry 40% of the final mark, and may keep '
                  'scoring while judging is locked.').font = NOTE
    row = write_header(ws, row + 2,
                       ['Name', 'Email', 'Type',
                        f"Teams scored (of {overview['total_teams']})"])
    for judge in exceptions:
        kind = 'External' if 'external' in str(judge.get('judge_type', '')).lower() else 'Internal'
        ws.cell(row=row, column=1, value=judge.get('name', ''))
        ws.cell(row=row, column=2, value=judge.get('email', ''))
        ws.cell(row=row, column=3, value=kind)
        ws.cell(row=row, column=4, value=judge_submitted_count(judge, stage_id))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='TEAMS WITH NO PANEL').font = H2
    ws.cell(row=row + 1, column=1,
            value='Only exception jury can score these.').font = NOTE
    row += 2
    if unassigned:
        row = write_header(ws, row, ['Team Code', 'Team Name', 'Leader'])
        for team in unassigned:
            ws.cell(row=row, column=1, value=team.get('team_code', ''))
            ws.cell(row=row, column=2, value=team.get('team_name', ''))
            ws.cell(row=row, column=3, value=team.get('leader_name', ''))
            row += 1
    else:
        ws.cell(row=row, column=1, value='(none — every team has a panel)')
        row += 1

    row += 1
    ws.cell(row=row, column=1,
            value=f"Totals: {overview['judges']} judges "
                  f"({overview['group']} group + {overview['exception']} exception) · "
                  f"{overview['total_teams']} teams "
                  f"({overview['assigned']} assigned, {overview['unassigned']} without a panel)"
            ).font = NOTE

    # ------------------------------------------------------------------ #
    # Panel 1..5
    # ------------------------------------------------------------------ #
    for p in panels:
        ws = wb.create_sheet(f"Panel {p['panel_no']}")
        ws.cell(row=1, column=1, value=f"Panel {p['panel_no']} — "
                f"{p['judge_count']} judges, {p['team_count']} teams, "
                f"{p['submitted']}/{p['expected']} evaluations ({p['pct']}%)").font = TITLE

        row = 3
        ws.cell(row=row, column=1, value=f"JUDGES ({p['judge_count']})").font = H2
        row = write_header(ws, row + 1,
                           ['Name', 'Email', 'Type', 'Coordinator',
                            f"Teams scored (of {p['team_count']})"],
                           widths=[28, 34, 10, 12, 22])
        if p['judges']:
            for judge in p['judges']:
                kind = 'External' if 'external' in str(judge.get('judge_type', '')).lower() else 'Internal'
                ws.cell(row=row, column=1, value=judge.get('name', ''))
                ws.cell(row=row, column=2, value=judge.get('email', ''))
                ws.cell(row=row, column=3, value=kind)
                ws.cell(row=row, column=4,
                        value='Yes' if judge.get('is_overall_jury_coordinator') else '')
                ws.cell(row=row, column=5, value=judge_submitted_count(judge, stage_id))
                row += 1
        else:
            ws.cell(row=row, column=1, value='(no judges assigned — these teams '
                                             'cannot be completed)')
            row += 1

        row += 1
        ws.cell(row=row, column=1, value=f"TEAMS ({p['team_count']})").font = H2
        row = write_header(ws, row + 1,
                           ['#', 'Team Code', 'Team Name', 'Leader', 'Leader Mobile',
                            'Leader Email', 'Panel Evals', 'Panel Expected',
                            'Exception Evals', 'Status'])
        if p['teams']:
            for index, team in enumerate(p['teams'], start=1):
                group_n, expected, exc_n, status = team_progress(team)
                ws.cell(row=row, column=1, value=index)
                ws.cell(row=row, column=2, value=team.get('team_code', ''))
                ws.cell(row=row, column=3, value=team.get('team_name', ''))
                ws.cell(row=row, column=4, value=team.get('leader_name', ''))
                ws.cell(row=row, column=5, value=team.get('leader_mobile', ''))
                ws.cell(row=row, column=6, value=team.get('leader_email', ''))
                ws.cell(row=row, column=7, value=group_n)
                ws.cell(row=row, column=8, value=expected)
                ws.cell(row=row, column=9, value=exc_n)
                ws.cell(row=row, column=10, value=status)
                row += 1
        else:
            ws.cell(row=row, column=1, value='(no teams assigned)')
            row += 1

    # ------------------------------------------------------------------ #
    # All teams (flat)
    # ------------------------------------------------------------------ #
    ws = wb.create_sheet('All teams')
    headers = ['Panel', 'Team Code', 'Team Name', 'Leader', 'Leader Mobile',
               'Leader Email', 'Panel Judges', 'Panel Evals', 'Panel Expected',
               'Exception Evals', 'Status', 'Registered At']
    write_header(ws, 1, headers,
                 widths=[12, 14, 26, 20, 14, 30, 60, 12, 14, 15, 20, 18])
    ws.freeze_panes = 'A2'

    # The panel roster repeats on every row of its panel on purpose: it makes
    # the flat sheet self-contained when sorted or filtered.
    roster_by_panel = {
        p['panel_no']: '; '.join(j.get('name', '') for j in p['judges'])
        for p in panels
    }
    all_docs = sorted(
        (team for p in panels for team in p['teams']),
        key=lambda t: (t.get('panel_no') or 0, str(t.get('team_name', '')).lower()),
    ) + unassigned

    row = 2
    for team in all_docs:
        group_n, expected, exc_n, status = team_progress(team)
        panel_no = team.get('panel_no')
        created = team.get('created_at')
        ws.cell(row=row, column=1,
                value=f'Panel {panel_no}' if panel_no else 'UNASSIGNED')
        ws.cell(row=row, column=2, value=team.get('team_code', ''))
        ws.cell(row=row, column=3, value=team.get('team_name', ''))
        ws.cell(row=row, column=4, value=team.get('leader_name', ''))
        ws.cell(row=row, column=5, value=team.get('leader_mobile', ''))
        ws.cell(row=row, column=6, value=team.get('leader_email', ''))
        ws.cell(row=row, column=7, value=roster_by_panel.get(panel_no, ''))
        ws.cell(row=row, column=8, value=group_n)
        ws.cell(row=row, column=9, value=expected)
        ws.cell(row=row, column=10, value=exc_n)
        ws.cell(row=row, column=11, value=status)
        ws.cell(row=row, column=12,
                value=created.strftime('%Y-%m-%d %H:%M') if created else '')
        row += 1

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# --------------------------------------------------------------------------- #
# Writes
# --------------------------------------------------------------------------- #

def _normalise_panel(panel_no):
    """None for 'no panel'; otherwise an int in 1..5. Raises on anything else."""
    if panel_no in (None, '', 'none', 'None'):
        return None
    panel = int(panel_no)
    if panel not in PANEL_NUMBERS:
        raise ValueError(f'panel_no must be 1..{PANEL_COUNT}, got {panel_no!r}')
    return panel


def set_team_panel(team_id, panel_no, actor_id=None):
    """Assign or clear one team's panel. Returns True if anything changed."""
    panel = _normalise_panel(panel_no)
    teams_col = get_teams_collection()

    try:
        oid = ObjectId(team_id)
    except Exception:
        return False

    team = teams_col.find_one({'_id': oid})
    if not team:
        return False

    before = team.get('panel_no')
    if before == panel:
        return False

    if panel is None:
        teams_col.update_one(
            {'_id': oid},
            {'$unset': {'panel_no': '', 'panel_assigned_at': '', 'panel_assigned_by': ''}},
        )
        log_audit(actor_id, 'team_panel_cleared', 'team', str(oid),
                  {'from': before, 'team_name': team.get('team_name')})
    else:
        teams_col.update_one({'_id': oid}, {'$set': {
            'panel_no': panel,
            'panel_assigned_at': datetime.utcnow(),
            'panel_assigned_by': actor_id,
        }})
        log_audit(actor_id, 'team_panel_updated', 'team', str(oid),
                  {'from': before, 'to': panel, 'team_name': team.get('team_name')})
    return True


def set_judge_panel(judge_id, panel_no, actor_id=None):
    """Move a judge onto a panel (group jury) or to the exception jury.

    `panel_no=None` means exception jury: scope flips to all_teams and panel_no
    is unset. Scope and membership are written together so the two can never
    disagree.
    """
    panel = _normalise_panel(panel_no)
    judges_col = get_judges_collection()

    try:
        oid = ObjectId(judge_id)
    except Exception:
        return False

    judge = judges_col.find_one({'_id': oid})
    if not judge:
        return False

    before_scope = judge.get('jury_scope')
    before_panel = judge.get('panel_no')

    if panel is None:
        if before_scope == SCOPE_ALL and before_panel is None:
            return False
        judges_col.update_one({'_id': oid}, {
            '$set': {'jury_scope': SCOPE_ALL, 'updated_at': datetime.utcnow()},
            '$unset': {'panel_no': ''},
        })
        log_audit(actor_id, 'judge_scope_updated', 'judge', str(oid), {
            'from': before_scope, 'to': SCOPE_ALL,
            'from_panel': before_panel, 'name': judge.get('name'),
        })
    else:
        if before_scope == SCOPE_ASSIGNED and before_panel == panel:
            return False
        judges_col.update_one({'_id': oid}, {'$set': {
            'jury_scope': SCOPE_ASSIGNED,
            'panel_no': panel,
            'updated_at': datetime.utcnow(),
        }})
        action = ('judge_scope_updated' if before_scope != SCOPE_ASSIGNED
                  else 'judge_panel_updated')
        log_audit(actor_id, action, 'judge', str(oid), {
            'from_scope': before_scope, 'to_scope': SCOPE_ASSIGNED,
            'from_panel': before_panel, 'to_panel': panel,
            'name': judge.get('name'),
        })
    return True


def auto_assign_teams(mode='blocks', overwrite=False, actor_id=None):
    """Distribute teams across the five panels.

    Sizes come from ceil(total / PANEL_COUNT), never a hardcoded 10 — teams are
    still registering, and a fixed 10 leaves the tail unassigned at any count
    that is not a multiple of five.

    overwrite=False (the default) touches only teams that have no panel, so the
    button is safe to press again after late registrations. Those late teams go
    to the panel holding the fewest teams rather than the next slot in sequence,
    which would otherwise pile them onto a panel that is already full.
    """
    teams_col = get_teams_collection()
    # _id as tie-break keeps the ordering stable across runs.
    teams = list(teams_col.find().sort([('created_at', 1), ('_id', 1)]))

    if overwrite:
        targets = teams
    else:
        targets = [t for t in teams if not t.get('panel_no')]

    per_panel = {n: 0 for n in PANEL_NUMBERS}
    if not overwrite:
        for t in teams:
            panel = t.get('panel_no')
            if panel in per_panel:
                per_panel[panel] += 1

    assigned = 0
    if overwrite:
        size = max(1, math.ceil(len(targets) / PANEL_COUNT))
        for index, team in enumerate(targets):
            if mode == 'round_robin':
                panel = (index % PANEL_COUNT) + 1
            else:
                panel = min(PANEL_COUNT, index // size + 1)
            if team.get('panel_no') != panel:
                teams_col.update_one({'_id': team['_id']}, {'$set': {
                    'panel_no': panel,
                    'panel_assigned_at': datetime.utcnow(),
                    'panel_assigned_by': actor_id,
                }})
                assigned += 1
            per_panel[panel] = per_panel.get(panel, 0) + 1
    else:
        for team in targets:
            # Smallest panel first, ties resolved by lowest panel number.
            panel = min(PANEL_NUMBERS, key=lambda n: (per_panel[n], n))
            teams_col.update_one({'_id': team['_id']}, {'$set': {
                'panel_no': panel,
                'panel_assigned_at': datetime.utcnow(),
                'panel_assigned_by': actor_id,
            }})
            per_panel[panel] += 1
            assigned += 1

    # per_panel is keyed by int, and BSON documents only accept string keys -
    # storing it raw raises InvalidDocument.
    log_audit(actor_id, 'team_panels_auto_assigned', 'jury_panel', mode, {
        'mode': mode, 'overwrite': overwrite,
        'assigned': assigned, 'skipped': len(teams) - len(targets),
        'per_panel': {str(k): v for k, v in per_panel.items()},
    })

    return {
        'assigned': assigned,
        'skipped': len(teams) - len(targets),
        'per_panel': per_panel,
    }


def clear_assignments(actor_id=None):
    """Remove every team's panel. Destructive; confirm before calling."""
    result = get_teams_collection().update_many(
        {'panel_no': {'$exists': True}},
        {'$unset': {'panel_no': '', 'panel_assigned_at': '', 'panel_assigned_by': ''}},
    )
    log_audit(actor_id, 'team_panels_reset', 'jury_panel', 'all',
              {'cleared': result.modified_count})
    return {'cleared': result.modified_count}
