"""Export judge roster (email/panel/scope) to Excel. Read-only.
Run with --db to also fill the 3 guest judges' stored temp_password from MongoDB."""
import sys, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv; load_dotenv()
from services.jury_roster import EXCEPTION_JURY, GROUP_JURY
UNIVERSAL = os.environ.get("JURY_UNIVERSAL_PASSWORD", "").strip() or "(not set – JURY_UNIVERSAL_PASSWORD)"

recs = {}
if '--db' in sys.argv:
    from pymongo import MongoClient
    db = MongoClient(os.environ['MONGO_URI'])[os.environ['MONGO_DB_NAME']]
    recs = {j['email'].lower(): j for j in db.judges.find({}, {'email':1,'temp_password':1,'credentials_sent':1})}

wb = Workbook(); ws = wb.active; ws.title = "Judges"
ws.append(["#", "Name", "Login Email", "Universal Password (any judge)", "Unique Password", "Panel", "Scope", "Judge Type", "Credentials Emailed"])
for c in ws[1]: c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1F4E78")

rows = [(j, "Exception jury (all teams, 40%)", "—") for j in EXCEPTION_JURY] + \
       [(j, f"Panel {j['panel_no']} (assigned teams, 60%)", j['panel_no']) for j in GROUP_JURY]
for i, (j, scope, panel) in enumerate(rows, 1):
    r = recs.get(j['email'].strip().lower(), {})
    internal = j.get('has_mailbox', True)
    pw = r.get('temp_password') or ("(hashed – not retrievable; judge uses emailed password)" if internal
                                    else "(stored in DB – run with --db or see Admin → Judges)")
    sent = ("Yes" if r.get('credentials_sent') else "No") if recs else ("n/a" if not internal else "see Admin → Judges")
    ws.append([i, j['name'], j['email'], UNIVERSAL, pw, panel, scope,
               j.get('judge_type', 'INTERNAL_JUDGE'), sent])
for i, w in enumerate([4, 26, 34, 30, 52, 7, 32, 16, 22], 1): ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
out = "TechForge_Judges_Credentials.xlsx"; wb.save(out); print("wrote", out, ws.max_row-1, "judges")
